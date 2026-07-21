"""CSV and XLSX export.

Exports are written to a temporary file while the file lock is held, then sent
as a download. Streaming straight from HDF5 to the socket would mean holding
that lock for as long as the client takes to read the response, which would
block every other request against the same file. Staging to disk keeps the lock
window tied to read speed instead of network speed.
"""

from __future__ import annotations

import csv
import tempfile
from pathlib import Path
from typing import Any

import pandas as pd

from . import jsonsafe
from .files import OpenFile
from .readers import parse_cols, parse_dim_slice
from .service import TIME_COLUMN_NAME, time_index_matches

CHUNK_ROWS = 50_000

# Excel's own hard limit is 1,048,576 rows including the header.
XLSX_ROW_LIMIT = 1_000_000


class ExportTooLargeError(Exception):
    pass


def _iter_chunks(
    open_file: OpenFile,
    path: str,
    *,
    start: int,
    stop: int | None,
    cols_spec: str | None,
    dim_slice_spec: str | None,
    use_time_index: bool,
):
    """Yield successive DataFrames covering the requested range."""
    reader = open_file.reader(path)
    if reader.decode_fallback:
        raise ValueError(reader.decode_fallback)

    start = max(0, start)
    stop = reader.nrows if stop is None else min(stop, reader.nrows)

    dim_slice = parse_dim_slice(dim_slice_spec, reader.ndim) if reader.ndim > 2 else None
    selection = parse_cols(cols_spec)

    apply_time = use_time_index and time_index_matches(open_file, reader)
    time_index = open_file.time_index() if apply_time else None

    for chunk_start in range(start, stop, CHUNK_ROWS):
        chunk_stop = min(chunk_start + CHUNK_ROWS, stop)
        frame = reader.read(chunk_start, chunk_stop, selection, dim_slice)
        if apply_time and time_index is not None:
            frame = frame.copy()
            frame.insert(0, TIME_COLUMN_NAME, list(time_index[chunk_start:chunk_stop]))
        yield frame


def export_rows(
    open_file: OpenFile,
    path: str,
    *,
    fmt: str = "csv",
    start: int = 0,
    stop: int | None = None,
    cols_spec: str | None = None,
    dim_slice_spec: str | None = None,
    use_time_index: bool = False,
) -> tuple[Path, str, str]:
    """Write the export and return (temp path, download filename, media type)."""
    reader = open_file.reader(path)
    total = (reader.nrows if stop is None else min(stop, reader.nrows)) - max(0, start)
    fmt = fmt.lower()

    leaf = reader.path.rstrip("/").rsplit("/", 1)[-1] or "data"
    stem = f"{open_file.path.stem}_{leaf}"

    kwargs = dict(
        start=start,
        stop=stop,
        cols_spec=cols_spec,
        dim_slice_spec=dim_slice_spec,
        use_time_index=use_time_index,
    )

    if fmt == "csv":
        target = _write_csv(open_file, path, **kwargs)
        return target, f"{stem}.csv", "text/csv"

    if fmt == "xlsx":
        if total > XLSX_ROW_LIMIT:
            raise ExportTooLargeError(
                f"{total:,} rows exceeds the {XLSX_ROW_LIMIT:,}-row XLSX limit "
                "(Excel itself cannot open more). Export as CSV instead, or "
                "narrow the row range."
            )
        target = _write_xlsx(open_file, path, **kwargs)
        return (
            target,
            f"{stem}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    raise ValueError(f"Unsupported export format {fmt!r}. Use csv or xlsx.")


def _write_csv(open_file: OpenFile, path: str, **kwargs: Any) -> Path:
    handle = tempfile.NamedTemporaryFile(
        mode="w", suffix=".csv", delete=False, newline="", encoding="utf-8"
    )
    target = Path(handle.name)
    try:
        with handle:
            writer = None
            for frame in _iter_chunks(open_file, path, **kwargs):
                if writer is None:
                    writer = csv.writer(handle)
                    writer.writerow(list(frame.columns))
                for row in jsonsafe.frame_to_rows(frame):
                    writer.writerow(["" if v is None else v for v in row])
    except Exception:
        target.unlink(missing_ok=True)
        raise
    return target


def _write_xlsx(open_file: OpenFile, path: str, **kwargs: Any) -> Path:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    handle.close()
    target = Path(handle.name)

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title="data")

    try:
        header_written = False
        date_columns: list[int] = []
        for frame in _iter_chunks(open_file, path, **kwargs):
            if not header_written:
                sheet.append(list(frame.columns))
                date_columns = [
                    i
                    for i, name in enumerate(frame.columns)
                    if pd.api.types.is_datetime64_any_dtype(frame[name])
                ]
                header_written = True

            # Keep real datetimes as datetimes so Excel formats them as dates
            # rather than as text, but pass NaT through as blank.
            records = frame.to_numpy(dtype=object)
            for row in records:
                cells = []
                for i, value in enumerate(row):
                    if i in date_columns:
                        cells.append(None if pd.isna(value) else pd.Timestamp(value).to_pydatetime())
                    else:
                        cells.append(jsonsafe.coerce_scalar(value))
                sheet.append(cells)

        if header_written:
            for i in date_columns:
                sheet.column_dimensions[get_column_letter(i + 1)].width = 12

        workbook.save(target)
    except Exception:
        target.unlink(missing_ok=True)
        raise
    finally:
        workbook.close()

    return target
